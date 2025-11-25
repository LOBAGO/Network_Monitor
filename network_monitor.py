import os
import subprocess
import time
import json
import datetime
from collections import defaultdict, deque
import threading
import requests
from functools import lru_cache
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import pychrome
import pyqtgraph as pg
from pyqtgraph.Qt import QtWidgets, QtCore, QtGui

import matplotlib.pyplot as plt
import matplotlib.dates as mdates

import sys

def get_base_path():
    """Return the directory where the script/exe is located."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


# ==================== Args ====================
BASE_DIR = get_base_path()
OUTPUT_FILE = os.path.join(BASE_DIR, "responses.jsonl")
ROLLING_SECONDS = 60
UPDATE_INTERVAL = 16
NUM_LINES = 3
FIXED_COLORS = ['#FF6B6B', "#FFC518", "#EAFA0F"]
MAX_RECORDS_PER_IP = 1000

CHROME_PATH = "C:/Program Files/Google/Chrome/Application/chrome.exe"
DEBUG_PORT = 9222
USER_DATA_DIR = "C:/ChromeDebug"

position = 0
record_data = defaultdict(lambda: deque(maxlen=MAX_RECORDS_PER_IP))
domain_record_data = defaultdict(lambda: deque(maxlen=MAX_RECORDS_PER_IP))  # Slot by domain
ip_to_isp_cache = {}
tab_listeners = {}
request_start_times = {}
request_ips = {}
request_domains = {}
is_monitoring = True
total_data_transferred = 0
session_start_time = datetime.datetime.now()

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    pass

def extract_domain(url):
    try:
        parsed = urlparse(url)
        hostname = parsed.hostname
        if not hostname:
            return "unknown"
        
        if hostname.startswith("www."):
            hostname = hostname[4:]
        
        return hostname
    except Exception:
        return "unknown"

def start_chrome():
    if not os.path.exists(USER_DATA_DIR):
        os.makedirs(USER_DATA_DIR)
    subprocess.Popen([
        CHROME_PATH,
        f'--remote-debugging-port={DEBUG_PORT}',
        f'--user-data-dir={USER_DATA_DIR}'
    ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

def attach_tab(tab):
    if tab.id in tab_listeners:
        return
    
    try:
        tab.start()
        tab.call_method("Network.enable")
        tab.call_method("Page.enable")

        def handle_request_will_be_sent(**kwargs):
            """ Capture Headers（Referer）"""
            request_id = kwargs.get("requestId")
            request = kwargs.get("request", {})
            timestamp = kwargs.get("timestamp")
            walltime = kwargs.get("walltime")

            request_start_times[request_id] = {
                'timestamp': timestamp,
                'walltime': walltime
            } 
            
            headers = request.get("headers", {})
            referer = headers.get("Referer") or headers.get("referer")
            
            if referer:
                domain = extract_domain(referer)
            else:
                url = request.get("url", "")
                domain = extract_domain(url)
            
            request_domains[request_id] = domain

        def handle_response_received(**kwargs):
            request_id = kwargs.get("requestId")
            response = kwargs.get("response", {})
            
            if response.get("fromDiskCache") or response.get("fromMemoryCache"):
                return
            
            ip = response.get("remoteIPAddress", "")
            request_ips[request_id] = ip

        def handle_loading_finished(**kwargs):
            global total_data_transferred
            try:
                request_id = kwargs.get("requestId")
                encoded_length = kwargs.get("encodedDataLength", 0)
                
                if encoded_length < 7*1000:
                    return
                
                start_info = request_start_times.get(request_id)
                if not start_info:
                    return
                
                start_timestamp = start_info['timestamp']
                end_timestamp = kwargs.get("timestamp")

                if end_timestamp is None or start_timestamp is None:
                    return
                

                # time lessthen 0.03 usually not real
                # TODO: consider tune the time or just ingore the data with time = 0.03
                # Python had a timestamp limit with 0.01 ~ 0.03 error, so a very small responce will had a very high error rate.
                duration = end_timestamp - start_timestamp
                duration = max(duration, 0.02)
                
                size_kb = encoded_length / 1000
                speed_mbps = encoded_length * 8 / (1000*1000) / duration
                ip = request_ips.get(request_id, "")
                domain = request_domains.get(request_id, "unknown")

                if domain == "unknown" or ip == "":
                    return
                
                # try to ingore all of the small responce
                # TODO: make sure if you really want to keep the small data or focus the method.
                if duration == 0.035:
                    return
                
                total_data_transferred += encoded_length

                wall_time = start_info.get('walltime')

                if wall_time:
                    actual_time = datetime.datetime.fromtimestamp(wall_time)
                    time_str = actual_time.strftime("%H:%M:%S")
                else:
                    time_str = time.strftime("%H:%M:%S", time.localtime())
                
                record = {
                    "time": time_str,
                    "size_kb": round(size_kb, 2),
                    "duration_s": round(duration, 3),
                    "speed_mbps": round(speed_mbps, 2),
                    "ip": ip,
                    "domain": domain,
                    "as": get_isp(ip)
                }
                
                with open(OUTPUT_FILE, "a", encoding="utf-8") as f:
                    f.write(json.dumps(record) + "\n")
                    f.flush()
                    
            except Exception as e:
                print(f"Fail to process response: {e}")

        tab.set_listener("Network.requestWillBeSent", handle_request_will_be_sent)
        tab.set_listener("Network.responseReceived", handle_response_received)
        tab.set_listener("Network.loadingFinished", handle_loading_finished)
        tab_listeners[tab.id] = tab
    except Exception as e:
        print(f"Fail to label: {e}")

def monitor_tabs():
    browser = pychrome.Browser(url=f"http://127.0.0.1:{DEBUG_PORT}")
    while is_monitoring:
        try:
            tabs = browser.list_tab()
            for tab in tabs:
                attach_tab(tab)
        except Exception:
            pass
        time.sleep(2)

@lru_cache(maxsize=256)
def get_isp(ip):
    if not ip or ip == "unknown":
        return "Unknown"
    
    try:
        url = f"https://ipinfo.io/{ip}/json"
        resp = requests.get(url, timeout=3)
        if resp.status_code == 200:
            data = resp.json()
            org = data.get("org", ip)
            if org.startswith("AS"):
                parts = org.split(" ", 1)
                if len(parts) > 1:
                    org = parts[1]
            return org
    except Exception as e:
        print(f"Fail to get isp {ip}: {e}")
    
    return ip

class SafeTimeAxis(pg.AxisItem):
    def tickStrings(self, values, scale, spacing):
        strs = []
        for v in values:
            try:
                if isinstance(v, (int, float)) and v > 0:
                    strs.append(datetime.datetime.fromtimestamp(v).strftime("%H:%M:%S"))
                else:
                    strs.append("")
            except Exception:
                strs.append("")
        return strs

class StatisticsPanel(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        
    def init_ui(self):
        layout = QtWidgets.QGridLayout(self)
        layout.setSpacing(10)
        
        self.labels = {}
        stats = [
            ("current_speed", "Current Speed", "0 Mbps"),
            ("peak_speed", "Peak Speed(Total in 1s)", "0 Mbps"),
            ("total_data", "Total Traffic", "0 MB"),
            ("active_ips", "Active IP", "0"),
            ("active_domains", "Active Domains", "0"),
            ("session_time", "Monitor Time", "00:00:00")
        ]
        
        for idx, (key, title, default) in enumerate(stats):
            row = idx // 3
            col = idx % 3
            
            frame = QtWidgets.QFrame()
            frame.setFrameStyle(QtWidgets.QFrame.StyledPanel)
            frame.setStyleSheet("""
                QFrame {
                    background-color: #2C3E50;
                    border-radius: 10px;
                    padding: 10px;
                }
            """)
            
            frame_layout = QtWidgets.QVBoxLayout(frame)
            frame_layout.setSpacing(5)
            
            title_label = QtWidgets.QLabel(title)
            title_label.setStyleSheet("color: #95A5A6; font-size: 30px;")
            title_label.setAlignment(QtCore.Qt.AlignCenter)
            
            value_label = QtWidgets.QLabel(default)
            value_label.setStyleSheet("color: #ECF0F1; font-size: 30px; font-weight: bold;")
            value_label.setAlignment(QtCore.Qt.AlignCenter)
            
            frame_layout.addWidget(title_label)
            frame_layout.addWidget(value_label)
            
            layout.addWidget(frame, row, col)
            self.labels[key] = value_label
    
    def update_stats(self, current_speed, peak_speed, total_mb, active_ips, active_domains):
        self.labels["current_speed"].setText(f"{current_speed:.1f} Mbps")
        self.labels["peak_speed"].setText(f"{peak_speed:.1f} Mbps")
        self.labels["total_data"].setText(f"{total_mb:.1f} MB")
        self.labels["active_ips"].setText(str(active_ips))
        self.labels["active_domains"].setText(str(active_domains))
        
        elapsed = datetime.datetime.now() - session_start_time
        hours, remainder = divmod(int(elapsed.total_seconds()), 3600)
        minutes, seconds = divmod(remainder, 60)
        self.labels["session_time"].setText(f"{hours:02d}:{minutes:02d}:{seconds:02d}")

class NetworkMonitorApp(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.position = 0
        self.peak_speed = 0
        self.line_labels_ip = [""] * NUM_LINES
        self.line_labels_domain = [""] * NUM_LINES
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("Network Traffic Monitor (IP & Domain)")
        self.setGeometry(100, 100, 1600, 900)
        self.setStyleSheet("""
            QWidget {
                background-color: #1E1E1E;
                color: #ECF0F1;
                font-family: 'Microsoft JhengHei', 'Segoe UI', Arial;
            }
            QPushButton {
                background-color: #3498DB;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-size: 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980B9;
            }
            QPushButton:pressed {
                background-color: #21618C;
            }
        """)
        
        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        title_layout = QtWidgets.QHBoxLayout()
        title_label = QtWidgets.QLabel("Network Traffic Monitor")
        title_label.setStyleSheet("font-size: 30px; font-weight: bold; color: #3498DB; font-family:Microsoft JhengHei;")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        
        self.btn_pause = QtWidgets.QPushButton("Pause")
        self.btn_pause.clicked.connect(self.toggle_monitoring)
        self.btn_export = QtWidgets.QPushButton("Export Plot")
        self.btn_export.clicked.connect(self.export_full_plot)
        self.btn_export_excel = QtWidgets.QPushButton("Export Excel")
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        self.btn_clear = QtWidgets.QPushButton("Clear Data")
        self.btn_clear.clicked.connect(self.clear_data)
        
        title_layout.addWidget(self.btn_pause)
        title_layout.addWidget(self.btn_export)
        title_layout.addWidget(self.btn_export_excel)
        title_layout.addWidget(self.btn_clear)
        
        main_layout.addLayout(title_layout)
        
        self.stats_panel = StatisticsPanel()
        main_layout.addWidget(self.stats_panel)
        
        charts_layout = QtWidgets.QHBoxLayout()
        
        time_axis_ip = SafeTimeAxis(orientation='bottom')
        self.plot_widget_ip = pg.GraphicsLayoutWidget()
        self.plot_widget_ip.setBackground('#2C3E50')
        
        self.plot_ip = self.plot_widget_ip.addPlot(
            title="<span style='color: #ECF0F1; font-size: 20px; font-family:Microsoft JhengHei;'>Traffic by IP/ISP</span>",
            axisItems={'bottom': time_axis_ip}
        )
        self.plot_ip.showGrid(x=True, y=True, alpha=0.5)
        self.plot_ip.setLabel('left', '<span style="color: #ECF0F1; font-family:Microsoft JhengHei;">Mbps</span>')
        self.plot_ip.setLabel('bottom', '<span style="color: #ECF0F1; font-family:Microsoft JhengHei;">Time</span>')
        self.plot_ip.addLegend(offset=(5, 5))
        
        self.lines_ip = []
        for idx, color in enumerate(FIXED_COLORS):
            line = self.plot_ip.plot([], [], pen=pg.mkPen(color, width=3), name="")
            self.lines_ip.append(line)
        
        charts_layout.addWidget(self.plot_widget_ip)
        
        time_axis_domain = SafeTimeAxis(orientation='bottom')
        self.plot_widget_domain = pg.GraphicsLayoutWidget()
        self.plot_widget_domain.setBackground('#2C3E50')
        
        self.plot_domain = self.plot_widget_domain.addPlot(
            title="<span style='color: #ECF0F1; font-size: 20px; font-family:Microsoft JhengHei;'>Traffic by Domain</span>",
            axisItems={'bottom': time_axis_domain}
        )
        self.plot_domain.showGrid(x=True, y=True, alpha=0.5)
        self.plot_domain.setLabel('left', '<span style="color: #ECF0F1; font-family:Microsoft JhengHei;">Mbps</span>')
        self.plot_domain.setLabel('bottom', '<span style="color: #ECF0F1; font-family:Microsoft JhengHei;">Time</span>')
        self.plot_domain.addLegend(offset=(5, 5))
        
        self.lines_domain = []
        for idx, color in enumerate(FIXED_COLORS):
            line = self.plot_domain.plot([], [], pen=pg.mkPen(color, width=3), name="")
            self.lines_domain.append(line)
        
        charts_layout.addWidget(self.plot_widget_domain)
        
        main_layout.addLayout(charts_layout)
        
        self.status_label = QtWidgets.QLabel("Monitoring...")
        self.status_label.setStyleSheet("color: #2ECC71; font-size: 20px; padding: 5px; font-family:Microsoft JhengHei;")
        main_layout.addWidget(self.status_label)
        
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.update_plot)
        self.timer.start(UPDATE_INTERVAL)
    
    def toggle_monitoring(self):
        if self.timer.isActive():
            self.timer.stop()
            self.btn_pause.setText("Continue")
            self.status_label.setText("Paused")
            self.status_label.setStyleSheet("color: #F39C12; font-size: 20px; padding: 5px; font-family:Microsoft JhengHei;")
        else:
            self.timer.start()
            self.btn_pause.setText("Pause")
            self.status_label.setText("Monitoring...")
            self.status_label.setStyleSheet("color: #2ECC71; font-size: 20px; padding: 5px; font-family:Microsoft JhengHei;")
    
    def clear_data(self):
        reply = QtWidgets.QMessageBox.question(
            self, 'Confirm', 'Are you sure to clear all data?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            global record_data, domain_record_data, total_data_transferred, session_start_time
            record_data.clear()
            domain_record_data.clear()
            total_data_transferred = 0
            session_start_time = datetime.datetime.now()
            self.position = 0
            self.peak_speed = 0
            
            with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
                pass
            
            QtWidgets.QMessageBox.information(self, 'Complete', 'Data cleared')
    
    def update_plot(self):
        now = datetime.datetime.now()
        
        try:
            with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
                f.seek(self.position)
                lines_read = f.readlines()
                self.position = f.tell()
        except FileNotFoundError:
            return
        
        for line in lines_read:
            try:
                record = json.loads(line.strip())
                ip = record.get("ip", "unknown")
                domain = record.get("domain", "unknown")
                dt = datetime.datetime.combine(
                    datetime.date.today(),
                    datetime.datetime.strptime(record["time"], "%H:%M:%S").time()
                )
                data_point = {"time": dt, "speed_mbps": record["speed_mbps"]}
                
                record_data[ip].append(data_point)
                
                if domain != "unknown":
                    domain_record_data[domain].append(data_point)
                    
            except Exception:
                continue
        
        if not record_data and not domain_record_data:
            return
        
        window_start = now - datetime.timedelta(seconds=ROLLING_SECONDS)
        
        self.update_chart(
            record_data, 
            self.lines_ip, 
            self.plot_ip, 
            self.line_labels_ip,
            window_start, 
            now,
            use_isp=True
        )
        
        self.update_chart(
            domain_record_data, 
            self.lines_domain, 
            self.plot_domain, 
            self.line_labels_domain,
            window_start, 
            now,
            use_isp=False
        )
        
        total_mb = total_data_transferred / (1024 * 1024)
        active_ips = len([ip for ip, records in record_data.items() if records])
        active_domains = len([d for d, records in domain_record_data.items() if records])
        
        current_total_speed = 0
        for records in list(record_data.values()) + list(domain_record_data.values()):
            if records:
                latest = records[-1]
                if latest["time"] >= now - datetime.timedelta(seconds=1):
                    current_total_speed += latest["speed_mbps"]
        
        self.peak_speed = max(self.peak_speed, current_total_speed)
        self.stats_panel.update_stats(current_total_speed, self.peak_speed, total_mb, active_ips, active_domains)
    
    def update_chart(self, data_dict, lines, plot, labels, window_start, now, use_isp=True):
        total_speeds = {
            key: sum(r["speed_mbps"] for r in records) 
            for key, records in data_dict.items()
        }
        top_keys = sorted(total_speeds, key=total_speeds.get, reverse=True)[:NUM_LINES]
        
        max_value = 1
        
        for idx in range(NUM_LINES):
            line = lines[idx]
            
            if idx < len(top_keys):
                key = top_keys[idx]
                
                data_dict[key] = deque(
                    [r for r in data_dict[key] if r["time"] >= window_start],
                    maxlen=MAX_RECORDS_PER_IP
                )
                
                per_second = defaultdict(float)
                for r in data_dict[key]:
                    t_sec = int(r["time"].timestamp())
                    per_second[t_sec] = max(per_second[t_sec], r["speed_mbps"])
                
                times = []
                values = []
                current = int(window_start.timestamp())
                end_time = int(now.timestamp())
                
                while current <= end_time:
                    times.append(current)
                    values.append(per_second.get(current, 0))
                    current += 1
                
                line.setData(times, values)
                
                if use_isp:
                    isp_name = get_isp(key)
                    label = f"{isp_name[:20]} {key}"
                else:
                    label = f"{key[:30]}"
                
                if labels[idx] != label:
                    plot.legend.items[idx][1].setText(label)
                    labels[idx] = label
                
                if values:
                    max_value = max(max_value, max(values) * 1.2)
            else:
                line.setData([], [])
                if labels[idx] != "":
                    labels[idx] = ""
        
        plot.setXRange(int(window_start.timestamp()), int(now.timestamp()))
        plot.setYRange(0, max_value)
    
    def export_full_plot(self):
        self.timer.stop()
        
        full_record_data = defaultdict(list)
        full_domain_data = defaultdict(list)
        
        try:
            with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
                for line in f:
                    try:
                        record = json.loads(line.strip())
                        ip = record.get("ip", "unknown")
                        domain = record.get("domain", "unknown")
                        dt = datetime.datetime.combine(
                            datetime.date.today(),
                            datetime.datetime.strptime(record["time"], "%H:%M:%S").time()
                        )
                        data_point = {"time": dt, "speed_mbps": record["speed_mbps"]}
                        full_record_data[ip].append(data_point)
                        if domain != "unknown":
                            full_domain_data[domain].append(data_point)
                    except Exception:
                        continue
        except FileNotFoundError:
            QtWidgets.QMessageBox.warning(self, 'Error', 'No data found')
            self.timer.start()
            return
        
        if not full_record_data and not full_domain_data:
            QtWidgets.QMessageBox.warning(self, 'Error', 'No data to export')
            self.timer.start()
            return
        
        plt.style.use('dark_background')
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 10))
        fig.patch.set_facecolor('#1E1E1E')
        
        self.plot_export_chart(full_record_data, ax1, "Traffic by IP/ISP", use_isp=True)
        
        self.plot_export_chart(full_domain_data, ax2, "Traffic by Domain", use_isp=False)
        
        plt.tight_layout()
        
        filename = os.path.join(BASE_DIR, f"network_traffic_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        plt.savefig(filename, dpi=150, facecolor='#1E1E1E')
        
        plt.show()
        QtWidgets.QMessageBox.information(self, 'Complete', f'Saved as: {filename}')
        self.timer.start()
    
    def plot_export_chart(self, data_dict, ax, title, use_isp=True):
        ax.set_facecolor('#2C3E50')
        
        total_speeds = {
            key: sum(r["speed_mbps"] for r in records) 
            for key, records in data_dict.items()
        }
        top_keys = sorted(total_speeds, key=total_speeds.get, reverse=True)[:NUM_LINES]
        
        for idx, key in enumerate(top_keys):
            records = data_dict[key]
            per_second = defaultdict(float)
            
            for r in records:
                t_sec = r["time"].replace(microsecond=0)
                per_second[t_sec] = max(per_second[t_sec], r["speed_mbps"])
            
            if not per_second:
                continue
            
            start_time = min(per_second.keys())
            end_time = max(per_second.keys())
            times = []
            values = []
            current = start_time
            
            while current <= end_time:
                times.append(current)
                values.append(per_second.get(current, 0))
                current += datetime.timedelta(seconds=1)
            
            if use_isp:
                isp_name = f"{get_isp(key)} ({key})"
                label = f"{isp_name}"
            else:
                label = f"{key[:25]}"
            
            ax.plot(times, values, 
                   label=label, 
                   color=FIXED_COLORS[idx % len(FIXED_COLORS)],
                   linewidth=2)
        
        ax.set_xlabel("Time", fontsize=12, color='#ECF0F1')
        ax.set_ylabel("Mbps", fontsize=12, color='#ECF0F1')
        ax.set_title(title, fontsize=14, fontweight='bold', color='#3498DB', pad=15)
        ax.legend(loc='upper left', framealpha=0.9, fontsize=9)
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))

    def export_to_excel(self):
        self.timer.stop()

        try:
            domain_records = defaultdict(list)
            with open(OUTPUT_FILE, "r", encoding= "utf-8") as f:
                for line in f:
                    try:
                        record = json.loads(line.strip())
                        domain =  record.get("domain","unknown")
                        if domain != "unknown":
                            domain_records[domain].append(record)
                    except Exception:
                        continue
            
            if not domain_records:
                QtWidgets.QMessageBox.warning(self,'Error' , 'No data to export')
                self.timer.start()
                return
            
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
        
            border_side = Side(style='thin', color="CCCCCC")
            border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
            cell_alignment = Alignment(horizontal="center", vertical="center")

            domain_totals = {domain: sum(r.get("size_kb", 0) for r in records) for domain, records in domain_records.items()}
            sorted_domains = sorted(domain_totals, key=domain_totals.get, reverse=True)
            for domain in sorted_domains:
                records = domain_records[domain]

                sheet_name = domain[:31]

                for char in ['\\', '/', '*', '?', ':', '[', ']']:
                    sheet_name = sheet_name.replace(char, '_')

                ws = wb.create_sheet(title=sheet_name)

                headers = ["Time", "Size (KB)", "Duration (s)", "Speed (Mbps)", "IP", "ISP/AS"]

                ws.append(headers)

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border

                for record in records:
                    row_data = [
                        record.get("time", ""),
                        record.get("size_kb", 0),
                        record.get("duration_s", 0),
                        record.get("speed_mbps", 0),
                        record.get("ip", ""),
                        record.get("as", "")
                    ]
                    ws.append(row_data)

                for row_num in range(2, len(records) + 2):
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.alignment = cell_alignment
                        cell.border = border

                        if col_num == 2:  # Size (KB)
                            cell.number_format = '#,##0.00'
                        elif col_num == 3:  # Duration (s)
                            cell.number_format = '0.000'
                        elif col_num == 4:  # Speed (Mbps)
                            cell.number_format = '#,##0.00'
                
                stats_row = len(records) + 3
                ws.cell(row=stats_row, column=1, value="Statistics:").font = Font(bold=True)
            
                total_size = sum(r.get("size_kb", 0) for r in records)
                avg_speed = sum(r.get("speed_mbps", 0) for r in records) / len(records) if records else 0
                max_speed = max((r.get("speed_mbps", 0) for r in records), default=0)
            
                ws.cell(row=stats_row + 1, column=1, value="Total Size (MB):")
                ws.cell(row=stats_row + 1, column=2, value=round(total_size / 1024, 2))
            
                ws.cell(row=stats_row + 2, column=1, value="Avg Speed (Mbps):")
                ws.cell(row=stats_row + 2, column=2, value=round(avg_speed, 2))
            
                ws.cell(row=stats_row + 3, column=1, value="Max Speed (Mbps):")
                ws.cell(row=stats_row + 3, column=2, value=round(max_speed, 2))
            
                ws.cell(row=stats_row + 4, column=1, value="Request Count:")
                ws.cell(row=stats_row + 4, column=2, value=len(records))
            
                for col_num in range(1, len(headers) + 1):
                    column_letter = get_column_letter(col_num)
                    max_length = 0
                    for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                        for cell in row:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                ws.freeze_panes = 'A2'

            summary_ws = wb.create_sheet(title="Summary", index=0)
            summary_headers = ["Domain", "Total Size (MB)", "Avg Speed (Mbps)", "Max Speed (Mbps)", "Request Count"]
            summary_ws.append(summary_headers)

            for col_num, header in enumerate(summary_headers, 1):
                cell = summary_ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            for domain in sorted_domains:
                records = domain_records[domain]
                total_size = sum(r.get("size_kb", 0) for r in records) / 1024
                avg_speed = sum(r.get("speed_mbps", 0) for r in records) / len(records) if records else 0
                max_speed = max((r.get("speed_mbps", 0) for r in records), default=0)
            
                row_data = [
                    domain,
                    round(total_size, 2),
                    round(avg_speed, 2),
                    round(max_speed, 2),
                    len(records)
                ]   
                summary_ws.append(row_data)

            for row_num in range(2, len(sorted_domains) + 2):
                for col_num in range(1, len(summary_headers) + 1):
                    cell = summary_ws.cell(row=row_num, column=col_num)
                    cell.alignment = cell_alignment
                    cell.border = border
                
                    if col_num in [2, 3, 4]:
                        cell.number_format = '#,##0.00'

            for col_num in range(1, len(summary_headers) + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
            
                for row in summary_ws.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                adjusted_width = min(max_length + 2, 50)
                summary_ws.column_dimensions[column_letter].width = adjusted_width

            summary_ws.freeze_panes = 'A2'
            filename = os.path.join(BASE_DIR, f"network_traffic_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(filename)

            QtWidgets.QMessageBox.information(self, 'Export Complete', f'Data exported successfully!\n\nFile: {filename}\n\nDomains: {len(sorted_domains)}\nTotal Records: {sum(len(r) for r in domain_records.values())}')

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Export Error', f'Failed to export: {str(e)}')
    
        finally:
            self.timer.start()

if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    app.setStyle('Fusion')
    
    threading.Thread(target=start_chrome, daemon=True).start()
    threading.Thread(target=monitor_tabs, daemon=True).start()
    
    window = NetworkMonitorApp()
    window.show()
    
    app.exec_()